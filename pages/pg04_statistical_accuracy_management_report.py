"""
統計学的精度管理　確認・台帳作成
"""

import calendar
from io import BytesIO

import sqlite3

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, Border, Side

import pandas as pd

import streamlit as st

st.sidebar.markdown('### 統計学的精度管理　確認・台帳作成')
st.markdown('## 統計学的精度管理　確認・台帳作成')

implementation_year = st.sidebar.text_input('西暦',
                                            placeholder='西暦をここに入力')
implementation_month = st.sidebar.text_input('月',
                                             placeholder='月（MM）をここに入力')

# 年と月が入力されている場合のみ日付を計算
if implementation_year and implementation_month:
    # 年月から最終日を算出
    _, last_day = calendar.monthrange(int(implementation_year),
                                      int(implementation_month))
    # 年+月+日（年月から最終日を算出）YYYY-MM-DD
    implementation_date = f"{
        implementation_year}-{implementation_month.zfill(2)}-{last_day}"

###############################################################################

    # データフレーム作成
    # 基本基本統計
    conn = sqlite3.connect('db_folder/qc_data_base.db')

    # SQLクエリを実行
    qc_data_query = f"""
    SELECT `date`, `batch`, `item_code`, `model`, `type`, `Lot Number`,
           `sd_conversion`, `Ct`, `judgment`, `Error_type`, `Violated_rule`
    FROM table_qc_mulch_rule
    WHERE `date` <= '{implementation_date}'
    ORDER BY `date` DESC
    """

    # クエリ結果をデータフレームに変換
    df_qc_query = pd.read_sql_query(qc_data_query, conn)

    conn.close()

    date_max = df_qc_query['Measurement Date'].max()
    date_min = df_qc_query['Measurement Date'].min()

    # QC_List 作成：データフレームを 'Measurement Date'と'Measurement Time'列でソート
    df_qc_query_data = df_qc_query.sort_values(['date', 'batch'],
                                               ascending=True)

    # 列名を定義
    column_names = pd.MultiIndex.from_product([['Ct', 'sd_conversion'],
                                               ['N', 'MEAN', 'STD', 'CV']])

    # Lot Numberでグループ化し、Ctとsd_conversionのN数、平均と標準偏差を算出
    grouped_mean_std_cv = round(
        df_qc_query_data.groupby('Lot Number')[['Ct', 'sd_conversion']].agg(
            ['count', 'mean', 'std', lambda x: x.std() / x.mean()]), 3)

    # 列名を変更
    grouped_mean_std_cv.columns = column_names

    # Lot Number毎にCTRL 1及びCTRL 2の工程能力指数を算出し、grouped_mean_std_cvに追加する
    grouped_mean_std_cv[('Ct', 'Cp')] = round(
        (df_qc_query_data['UCL 1'].mean() - df_qc_query_data['LCL 1'].mean()) /
        (grouped_mean_std_cv[('Ct', 'STD')] * 6), 3)
    grouped_mean_std_cv[('sd_conversion', 'Cp')] = round(
        (df_qc_query_data['UCL 2'].mean() - df_qc_query_data['LCL 2'].mean()) /
        (grouped_mean_std_cv[('sd_conversion', 'STD')] * 6), 3)

    # DataFrame：basic_statistic:基本統計量
    df_basic_statistic = grouped_mean_std_cv

    # 管理試料結果
    df_qc_query_data_list = df_qc_query_data.drop(columns=['Ct',
                                                           'sd_conversion'])

    # DataFrame：df_qc_query_data_list:管理試料測定結果
    df_qc_query_data_list = df_qc_query_data_list.rename(
        columns={
            'date': '実施日',
            'batch': 'Batch',
            'item_code': 'Item Code',
            'model': 'Model',
            'type': 'Type',
            'Lot Number': 'Lot Number',
            'QC Results': 'Multi-Rule',
            'Cont 1': 'CTRL1(SDc)',
            'CTRL 1': 'CTRL 1',
            'Cont 2': 'CTRL2(SDc)',
            'CTRL 2': 'CTRL 2'
        })

############################################################################

    # 異常判定対応記録
    conn = sqlite3.connect('Control_Data.db')

    # SQLクエリを実行
    qc_act_query_list = f"""
    SELECT `date`, `batch`, `item_code`, `model`, `type`, `Lot Number`,
           `QC_Results`, `cause`, `CAPA`
    FROM table_qc_act_log
    WHERE `date` <= '{date_max}'
    AND `date` >= '{date_min}'
    AND  `CAPA` != 'None'
    ORDER BY `date` DESC
    """

    # クエリ結果をデータフレームに変換
    df_qc_act_query_list = pd.read_sql_query(
        qc_act_query_list, conn).reset_index(drop=True)
    df_qc_act_query_list = df_qc_act_query_list.sort_values(
        ['date', 'batch'], ascending=True)

    # データが存在する場合
    if len(df_qc_act_query_list) > 0:
        # 'Measurement Date'列のフォーマットを'MM-DD'に変更
        df_qc_act_query_list['date'] = pd.to_datetime(
            df_qc_act_query_list['date']).dt.strftime('%Y-%m-%d')

        # 'Measurement Time'列のフォーマットを'HH:MM'に変更
        df_qc_act_query_list['batch'] = pd.to_datetime(
            df_qc_act_query_list['batch']).dt.strftime('%H:%M')

    conn.close()

    # Column Rename
    df_qc_act_query_list = df_qc_act_query_list.rename(
        columns={
            'date': '実施日',
            'batch': '開始時間',
            'QC_Results': 'Multi-Rule判定',
            'cause': '推測される原因',
            'CAPA': '予防処置・是正処置'
        })

##############################################################################

    # 測定実施状況
    conn = sqlite3.connect('Control_Data.db')

    # SQLクエリを実行
    qc_situation_query = f"""
    SELECT `date`, `batch`, `standard_usage_lot`, `standard_usage_open`,\
           `control_usage_lot`, `control_usage_open`, `reagents_usage_lot`, \
           `reagents_usage_open`, `measuring_instrument_usage`, \
           `measuring_instrument_change`
    FROM table_qc_act_log
    WHERE `Measurement Date` <= '{date_max}'
    AND `Measurement Date` >= '{date_min}'
    AND `measuring_instrument_change` != ''
    ORDER BY `Measurement Date` DESC
    """

    # クエリ結果をデータフレームに変換
    df_qc_situation_query = pd.read_sql_query(
        qc_situation_query, conn).reset_index(drop=True)
    # データフレームを'Measurement Date'と'Measurement Time'列でソート
    df_qc_situation_query = df_qc_situation_query.sort_values(
        ['date', 'batch'], ascending=True)

    if len(df_qc_situation_query) > 0:
        # 'Measurement Date'列のフォーマットを'MM-DD'に変更
        df_qc_situation_query['date'] = pd.to_datetime(
            df_qc_situation_query['date']).dt.strftime('%Y-%m-%d')

        # 'Measurement Time'列のフォーマットを数値に変更
        df_qc_situation_query['batch'] = pd.to_numeric(
            df_qc_situation_query['batch'])

    conn.close()

    columns_to_convert = ['standard_usage_lot',
                          'standard_usage_open',
                          'control_usage_lot',
                          'control_usage_open',
                          'reagents_usage_lot',
                          'reagents_usage_open',
                          'measuring_instrument_usage',
                          'measuring_instrument_change']

    df_qc_situation_query[columns_to_convert] = df_qc_situation_query[
        columns_to_convert].replace({1: 'あり', 0: '－'})

    # Column Rename
    df_qc_situation_query = df_qc_situation_query.rename(
        columns={
            'date': '実施日',
            'batch': '開始時間',
            'standard_usage_lot': 'STD:Lot変更',
            'standard_usage_open': 'STD:新規開封',
            'control_usage_lot': 'CTRL:Lot変更',
            'control_usage_open': 'CTRL:新規開封',
            'reagents_usage_lot': '検査薬:Lot変更',
            'reagents_usage_open': '検査薬:新規開封',
            'measuring_instrument_usage': '機器異常',
            'measuring_instrument_change': '部品交換'
        })

    df_qc_situation_query = df_qc_situation_query.T

##############################################################################

    # SQLite3に接続しデータの追加
    conn = sqlite3.connect('db_folder/qc_data_base.db')

    # SQLクエリを実行
    qc_check_od0_query = f"""
        SELECT `date`, `batch`, `Lot Number`, `SD Conversion`, `Measurements`
        FROM table_qc_data
        WHERE `QC Code` = "STD1"
        AND `date` <= '{date_max}'
        AND `date` >= '{date_min}'
        ORDER BY `date` DESC
        """

    qc_check_od1_query = f"""
        SELECT `date`, `batch`, `Lot Number`, `SD Conversion`, `Measurements`
        FROM table_qc_data
        WHERE `QC Code` = "STD2"
        AND `date` <= '{date_max}'
        AND `date` >= '{date_min}'
        ORDER BY `date` DESC
        """

    qc_check_od2_query = f"""
        SELECT `date`, `batch`, `Lot Number`, `SD Conversion`, `Measurements`
        FROM table_qc_data
        WHERE `QC Code` = "STD3"
        AND `date` <= '{date_max}'
        AND `date` >= '{date_min}'
        ORDER BY `date` DESC
        """

    qc_check_od3_query = f"""
        SELECT `date`, `batch`, `Lot Number`, `SD Conversion`, `Measurements`
        FROM table_qc_data
        WHERE `QC Code` = "STD4"
        AND `date` <= '{date_max}'
        AND `date` >= '{date_min}'
        ORDER BY `date` DESC
        """

    qc_check_od4_query = f"""
        SELECT `date`, `batch`, `Lot Number`, `SD Conversion`, `Measurements`
        FROM table_qc_data
        WHERE `QC Code` = "STD5"
        AND `date` <= '{date_max}'
        AND `date` >= '{date_min}'
        ORDER BY `date` DESC
        """

    qc_check_od_blank_query = f"""
        SELECT `date`, `batch`, `Lot Number`, `SD Conversion`, `Measurements`
        FROM table_qc_data
        WHERE `QC Code` = "BLANK"
        AND `date` <= '{date_max}'
        AND `date` >= '{date_min}'
        ORDER BY `date` DESC
        """

    df_qc_od0_query = pd.read_sql_query(qc_check_od0_query, conn)
    df_qc_od1_query = pd.read_sql_query(qc_check_od1_query, conn)
    df_qc_od2_query = pd.read_sql_query(qc_check_od2_query, conn)
    df_qc_od3_query = pd.read_sql_query(qc_check_od3_query, conn)
    df_qc_od4_query = pd.read_sql_query(qc_check_od4_query, conn)
    df_qc_od_blank_query = pd.read_sql_query(qc_check_od_blank_query, conn)

    df_qc_od0_query = df_qc_od0_query.rename(columns={
        'SD Conversion': 'STD1 SDc',
        'Measurements': 'STD1'
    })

    df_qc_od1_query = df_qc_od1_query.rename(columns={
        'SD Conversion': 'STD2 SDc',
        'Measurements': 'STD2'
    })

    df_qc_od2_query = df_qc_od2_query.rename(columns={
        'SD Conversion': 'STD3 SDc',
        'Measurements': 'STD3'
    })

    df_qc_od3_query = df_qc_od3_query.rename(columns={
        'SD Conversion': 'STD4 SDc',
        'Measurements': 'STD4'
    })

    df_qc_od4_query = df_qc_od4_query.rename(columns={
        'SD Conversion': 'STD5 SDc',
        "Measurements": "STD5",
    })

    df_qc_od_blank_query = df_qc_od_blank_query.rename(columns={
        'SD Conversion': 'BLANK SDc',
        'Measurements': 'BLANK'
    })

    df_qc_od_query_data = pd.merge(
        df_qc_od0_query, df_qc_od1_query, on=['date', 'batch', 'Lot Number'])
    df_qc_od_query_data = pd.merge(
        df_qc_od_query_data,
        df_qc_od2_query,
        on=['date', 'batch', 'Lot Number'])
    df_qc_od_query_data = pd.merge(
        df_qc_od_query_data,
        df_qc_od3_query,
        on=['date', 'batch', 'Lot Number'])
    df_qc_od_query_data = pd.merge(
        df_qc_od_query_data,
        df_qc_od4_query,
        on=['date', 'batch', 'Lot Number'])
    df_qc_od_query_data = pd.merge(
        df_qc_od_query_data,
        df_qc_od_blank_query, on=['date', 'batch', 'Lot Number'])

    df_qc_od_query_data = df_qc_od_query_data.sort_values(
        ['date', 'batch'], ascending=True)
    df_qc_od_query_data = df_qc_od_query_data[
        ['date', 'batch', 'STD1 SDc', 'STD2 SDc', 'STD3 SDc', 'STD4 SDc',
         'STD5 SDc', 'BLANK SDc']]

##############################################################################

    # 画面表示
    st.markdown('統計学的精度管理 確認')

    st.write('基本統計量', grouped_mean_std_cv)
    st.write('管理試料測定結果', df_qc_query_data_list)
    st.write('異常判定対応記録', df_qc_act_query_list)
    st.write('測定実施状況', df_qc_situation_query)

    st.write('x管理図(SD換算)濃度')
    st.line_chart(df_qc_query_data, use_container_width=True,
                  x='date', y=['Cont 1', 'Cont 2'])
    st.write('吸光度（OD値）一覧', df_qc_od_query_data)
    st.write('x管理図(SD換算) 吸光度（OD値）')
    st.line_chart(df_qc_od_query_data, use_container_width=True,
                  x='date',
                  y=['STD1 SDc', 'STD2 SDc', 'STD3 SDc', 'STD4 SDc',
                     'STD5 SDc', 'BLANK SDc'])

##############################################################################

    # Excelファイルを作成
    wb = Workbook()
    ws = wb.active

    # メモリ上にExcelを保存するためのバッファを作成
    excel_buffer = BytesIO()

    # pandasのExcelWriterを使用してExcelファイルを作成する
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        # データフレームを指定した位置に書き込む
        df_basic_statistic.to_excel(writer,
                                    sheet_name='Sheet1',
                                    startrow=25, startcol=1)  # B20から開始
        df_qc_query_data_list.to_excel(writer,
                                       sheet_name='Sheet1',
                                       startrow=7)           # A8から開始
        df_qc_act_query_list.to_excel(writer,
                                      sheet_name='Sheet1',
                                      startrow=32)            # A33から開始
        df_qc_situation_query.to_excel(writer,
                                       sheet_name='Sheet1',
                                       startrow=38, startcol=1)  # B39から開始

        # writerオブジェクトからワークブックを取得
        wb = writer.book
        ws = wb['Sheet1']

        # df_qc_query_data_listのデータを書き込んだ後の行数を取得
        max_row = ws.max_row

        # 'CTRL1(SDc)'のデータがF列、'CTRL2(SDc)'のデータがH列にある
        data_ctrl1 = Reference(ws,
                               min_col=6,
                               min_row=8,
                               max_col=6,
                               max_row=23)  # 'CTRL1(SDc)'の範囲(F8:F23)
        data_ctrl2 = Reference(ws,
                               min_col=8,
                               min_row=8,
                               max_col=8,
                               max_row=23)  # 'CTRL2(SDc)'の範囲(H8:H23)

        # '実施日'と'開始時間'はB列とC列に結合されています
        categories = Reference(ws,
                               min_col=2,
                               min_row=9,
                               max_row=23)  # B9:C23の範囲

        # LineChartオブジェクトを作成
        chart = LineChart()
        # CTRL1(SDc)のデータを追加
        chart.add_data(data_ctrl1, titles_from_data=True)
        # CTRL2(SDc)のデータを追加
        chart.add_data(data_ctrl2, titles_from_data=True)
        # カテゴリ（x軸のデータ）を設定
        chart.set_categories(categories)

        # グラフの幅を1.5倍に設定
        chart.width = int(chart.width * 1.2)

        # y軸の表示範囲を-4～4に設定
        chart.y_axis.scaling.min = -4
        chart.y_axis.scaling.max = 4
        # グラフのタイトルを設定
        chart.title = "x管理図（SD Conversion）"
        # グラフをシートの特定の位置（J8セル）に挿入
        ws.add_chart(chart)

############################################################################
        # セルB1のフォントを太字に設定し、そこにタイトルを書く。
        font = Font(bold=True, size=16)
        ws['B1'].font = font
        ws['B1'] = "統計学的精度管理台帳"

        # セルにラベルを書き込む
        font_title = Font(bold=True, size=12)
        ws['B2'] = "項目"
        ws['C2'].font = font_title
        ws['C2'] = "：Zonulin"
        ws['B3'] = "control"
        ws['C3'].font = font_title
        ws['C3'] = "：Kit付属"
        ws['B4'] = "期間"
        ws['C4'].font = font_title
        ws['C4'] = f"：{implementation_year}年{implementation_month}月分"
        ws['B5'] = "管理手法"
        ws['C5'].font = font_title
        ws['C5'] = "：Multi-Rule"
        ws['N1'] = "管理者"
        ws['P1'] = "　　　精度管理責任者\n　\n　"   # Top/Center表示ができなかったため、強制
        ws['P1'].alignment = Alignment(wrap_text=True)
        ws['P1'] = "染色体及び遺伝子検査に関する精度保証責任者"
        ws['L5'] = "／"
        ws['N5'] = "／"
        ws['P5'] = "／"

        ws['B25'].font = font_title
        ws['B25'] = '基本統計量'
        ws['B26'] = 'Lot Number'
        ws['B7'].font = font_title
        ws['B7'] = '管理試料測定結果'
        ws['J7'].font = font_title
        ws['J7'] = 'x管理図（SD Conversion）'
        ws['B32'].font = font_title
        ws['B32'] = '予防処置・是正処置'
        ws['B38'].font = font_title
        ws['B38'] = '実施状況'

        # 罫線の定義
        thin_border_side = Side(border_style="thin", color="000000")
        no_border_side = Side(border_style=None)

        # 文字位置の設定
        alignment_top_center = Alignment(vertical='top', horizontal='center')
        alignment_center = Alignment(vertical='center', horizontal='center')

        ws['B1'].alignment = Alignment(vertical='top')

        # 結合を解除してから再度結合
        for merged_cell in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged_cell.bounds
            ws.unmerge_cells(start_row=min_row,
                             start_column=min_col,
                             end_row=max_row,
                             end_column=max_col)

            for row in ws.iter_rows(min_row=min_row,
                                    max_row=max_row,
                                    min_col=min_col,
                                    max_col=max_col):
                for cell in row:
                    if min_row == max_row:  # 上下に結合されたセル
                        cell.alignment = alignment_top_center
                    else:  # 左右に結合されたセル
                        cell.alignment = alignment_center

            ws.merge_cells(start_row=min_row,
                           start_column=min_col,
                           end_row=max_row,
                           end_column=max_col)

        # 最後に特定のセルを結合
        ws.merge_cells('N1:O1')
        ws.merge_cells('P1:Q1')
        ws.merge_cells('N5:O5')
        ws.merge_cells('P5:Q5')

        # 結合したセルの配置を設定
        ws['N1'].alignment = alignment_top_center
        ws['O1'].alignment = alignment_top_center
        ws['P1'].alignment = alignment_top_center
        ws['Q1'].alignment = alignment_top_center
        ws['N5'].alignment = alignment_top_center
        ws['O5'].alignment = alignment_top_center
        ws['P5'].alignment = alignment_top_center
        ws['Q5'].alignment = alignment_top_center

        # 罫線を適用するセル範囲
        for row in ws.iter_rows(min_row=1,
                                max_row=5,
                                min_col=14,
                                max_col=17):
            for cell in row:
                top_side = thin_border_side if cell.row == 1 \
                    else no_border_side
                bottom_side = thin_border_side if cell.row == 5 \
                    else no_border_side
                left_side = thin_border_side if cell.column in [14, 16] \
                    else no_border_side
                right_side = thin_border_side if cell.column == 17 \
                    else no_border_side

                cell.border = Border(top=top_side,
                                     left=left_side,
                                     right=right_side,
                                     bottom=bottom_side)

        # 罫線の定義
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # 罫線を適用するセル範囲（格子状）
        for row in ws['B26:L30']:
            for cell in row:
                cell.border = thin_border

        for row in ws['A8:I23']:
            for cell in row:
                cell.border = thin_border

        for row in ws['B33:Q36']:
            for cell in row:
                cell.border = thin_border

        for row in ws['B39:Q49']:
            for cell in row:
                cell.border = thin_border

        # F33:F36をH33:H36に移動
        ws.move_range("F33:F36", rows=0, cols=2)
        # E33:E36をF33:F36に移動
        ws.move_range("E33:E36", rows=0, cols=1)

        # 結合するセル範囲のリスト
        merge_ranges = ['D33:E33', 'F33:G33', 'H33:Q33',
                        'D34:E34', 'F34:G34', 'H34:Q34',
                        'D35:E35', 'F35:G35', 'H35:Q35',
                        'D36:E36', 'F36:G36', 'H36:Q36']

        # 各範囲に対して結合操作を実行
        for merge_range in merge_ranges:
            ws.merge_cells(merge_range)

        # 列の幅を設定
        # セルの幅を設定
        column_widths = {'A': 0.1, 'B': 17, 'C': 12, 'D': 12, 'E': 12, 'F': 12,
                         'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12, 'L': 12,
                         'M': 12, 'N': 12, 'O': 12, 'P': 12, 'Q': 12}

        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

        # A1の行の高さを40に設定
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[28].height = 0.1
        ws.row_dimensions[39].height = 0.1

        # P1のセルを折り返し全体表示に設定
        ws['P1'].alignment = Alignment(wrap_text=True)

        # 印刷設定を変更
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # 印刷向きを横向きに設定
        ws.page_setup.fitToPage = True  # 内容を1ページにまとめる
        ws.page_setup.fitToHeight = 0  # 高さ方向には拡大/縮小しない
        ws.page_setup.fitToWidth = 1  # 幅方向には必要に応じて拡大/縮小する

###############################################################################

    # メモリバッファのポインタを先頭に戻す
    excel_buffer.seek(0)

    # Streamlitのダウンロードボタンを作成
    st.download_button(
        label='統計学的精度管理台帳・ダウンロード',
        data=excel_buffer.getvalue(),  # BytesIOオブジェクトからバイトデータを取得
        file_name=f'QC_{implementation_year}' +\
                  f'{implementation_month}_zonulin.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.\
            sheet'
    )

    st.markdown("※ ボタンをクリックすると、ダウンロードが開始されます。")
