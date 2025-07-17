"""
# -*- coding: utf-8 -*-
# 統計学的精度管理　確認・台帳作成
"""

import calendar
import sqlite3

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows


# メインの処理を行う関数
def create_excel_report(df, year, month):
    """エクセルファイルを作成する関数"""
    # ファイル名を生成
    filename = f"内部精度管理台帳_{year}年{month.zfill(2)}月.xlsx"

    # 新しいワークブックを作成
    wb = Workbook()

    # デフォルトシートを削除
    if wb.active:
        wb.remove(wb.active)

    # データシートを作成
    ws_data = wb.create_sheet("データ")

    # データフレームをエクセルに書き込み
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)

    # グラフシートを作成
    ws_chart = wb.create_sheet("グラフ")

    # SD換算データがある場合、グラフを作成
    if not df.empty and "PC/SD換算" in df.columns and "Type" in df.columns:
        # データを日時でソート
        df_sorted = df.sort_values("日時")
        # 日付ごと・Typeごとにピボット
        pivot = df_sorted.pivot_table(index="日時", columns="Type", values="PC/SD換算")
        pivot = pivot.reset_index()
        # グラフデータをエクセルに書き込み（A1から開始）
        start_row = 1
        for r in dataframe_to_rows(pivot, index=False, header=True):
            ws_chart.append(r)

        # 折れ線グラフを作成（画面表示と同じスタイル）
        if not pivot.empty:
            chart = LineChart()
            chart.title = f"{year}年{month}月 SD換算推移（Typeごと）"
            chart.x_axis.title = "日時"
            chart.y_axis.title = "SD換算"
            chart.y_axis.scaling.min = -5
            chart.y_axis.scaling.max = 5

            # グラフのサイズを設定
            chart.width = 15  # 幅を設定
            chart.height = 10  # 高さを設定

            # データ範囲を正確に設定（ヘッダー行を含む）
            data_start_row = start_row  # ヘッダー行から開始
            data_end_row = start_row + len(pivot)
            data_start_col = 2  # B列から（A列は日付）
            data_end_col = 1 + len(pivot.columns) - 1  # Typeの数分

            # データ範囲（B列以降のSD換算値、ヘッダー行を含む）
            data_range = Reference(
                ws_chart,
                min_col=data_start_col,
                max_col=data_end_col,
                min_row=data_start_row,
                max_row=data_end_row,
            )
            # カテゴリ範囲（A列の日付、ヘッダー行を除く）
            categories = Reference(
                ws_chart, min_col=1, min_row=data_start_row + 1, max_row=data_end_row
            )

            # カラム名（Type名）を件名として表示
            chart.add_data(data_range, titles_from_data=True)
            chart.set_categories(categories)

            # グラフをA1に配置
            ws_chart.add_chart(chart, "A1")

    # ファイルを保存
    wb.save(filename)
    return filename


def main():
    """メイン関数"""
    # ページ設定
    st.set_page_config(page_title="精度管理図出力", layout="wide")
    # サイドバーとメイン画面にタイトルを表示
    st.sidebar.markdown("### 統計学的精度管理　確認・台帳作成")
    st.markdown("## 統計学的精度管理　確認・台帳作成")

    # サイドバーで年月を入力
    year = st.sidebar.text_input("西暦", placeholder="西暦をここに入力")
    month = st.sidebar.text_input("月", placeholder="月（MM）をここに入力")

    # データ表示ボタンを追加
    show_data_button = st.sidebar.button("データ表示")

    # 内部精度管理台帳出力ボタンを追加
    export_excel_button = st.sidebar.button("内部精度管理台帳出力")

    # 年月とボタンが両方入力された場合の処理
    if show_data_button and year and month:
        _, last_day = calendar.monthrange(int(year), int(month))
        start_date = f"{year}-{month.zfill(2)}-01"
        end_date = f"{year}-{month.zfill(2)}-{last_day}"

        with sqlite3.connect("db_folder/qc_data_base.db", timeout=30) as conn:
            query = """
                SELECT
                  f.File_Name,
                  f.Date_Time AS '日時',
                  f.Batch AS 'バッチ',
                  f.Item_Code AS "項目コード",
                  f.Model AS "機種",
                  n.Type AS "Type",
                  n.Dye AS "Dye",
                  n.Lot_Number AS "Lot Number",
                  n.Verdict AS "NC判定",
                  p.Ct AS "PC/Ct値",
                  p.SD_Conversion AS "PC/SD換算",
                  m.judgment AS "PC/判定",
                  m.Violated_rule AS "違反ルール",
                  a.Cause AS "原因",
                  a.CAPA AS "CAPA"
                FROM table_qc_file_info f
                LEFT JOIN table_qc_nc n
                  ON f.File_Name = n.File_Name
                LEFT JOIN table_qc_pc p
                  ON n.File_Name = p.File_Name AND n.Dye = p.Dye
                LEFT JOIN table_qc_multi_rule m
                  ON p.File_Name = m.File_Name AND p.Type = m.Type
                LEFT JOIN table_qc_status_log s
                  ON f.File_Name = s.File_Name
                LEFT JOIN table_qc_act_log a
                  ON p.File_Name = a.File_Name AND p.Type = a.Type
                WHERE
                  strftime('%Y-%m-%d', replace(f.Date_Time, '/', '-')) >= ?
                  AND strftime('%Y-%m-%d', replace(f.Date_Time, '/', '-')) <= ?
                ORDER BY
                  f.Item_Code,
                  n.Type,
                  f.Date_Time DESC
                """
            df = pd.read_sql_query(query, conn, params=(start_date, end_date))
            df = df.drop_duplicates()

        # measurement_item.xlsxを読み込み、項目コードで結合
        df_item = pd.read_excel("data/master/measurement_item.xlsx")
        df = pd.merge(
            df, df_item, left_on="項目コード", right_on="Item_Code", how="left"
        )

        # SD換算のグラフをTypeごとにまとめて1つのグラフで表示
        if not df.empty and "PC/SD換算" in df.columns and "Type" in df.columns:
            st.markdown(f"### {year}年{month}月：ｘ管理図（SD換算）")
            df_sorted = df.sort_values("日時")
            # 日付表示を「YY/MM/DD：Batch」に変換
            df_sorted["日付バッチ"] = (
                pd.to_datetime(df_sorted["日時"]).dt.strftime("%y/%m/%d")
                + ":"
                + df_sorted["バッチ"].astype(str)
            )
            fig = px.line(
                df_sorted,
                x="日付バッチ",
                y="PC/SD換算",
                color="Type",
                # title="SD換算の推移（Typeごと）",
            )
            fig.update_yaxes(range=[-5, 5])
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("SD換算データがありません。")

        # QCチェックログの期間に該当するレコードを別テーブルで表示
        with sqlite3.connect("db_folder/qc_data_base.db", timeout=30) as conn:
            check_log_query = (
                "SELECT "
                "  check_date AS '確認日', "
                "  measurement AS '責任区分', "
                "  measurer AS '担当者名', "
                "  Type, "
                "  check_result AS 'チェック', "
                "  comment AS 'コメント' "
                "FROM table_qc_check_log "
                "WHERE "
                "  (strftime('%Y-%m-%d', start_date) <= ? "
                "   AND strftime('%Y-%m-%d', end_date) >= ?) "
                "ORDER BY start_date DESC"
            )
            df_check_log = pd.read_sql_query(
                check_log_query, conn, params=(start_date, end_date)
            )

        st.markdown("### QCチェックログ")
        if not df_check_log.empty:
            grouped = df_check_log.groupby("Type")
            for type_val, group in grouped:
                st.markdown(f"#### Type: {type_val}")
                st.dataframe(group, use_container_width=True)
        else:
            st.info("該当期間のQCチェックログはありません。")

        st.markdown(f"### {year}年{month}月：データ一覧")
        if not df.empty:
            grouped = df.groupby(["Measurement_Item", "Type"])
            for (item_name, type_val), group in grouped:
                st.markdown(f"#### 項目名: {item_name} / Type: {type_val}")
                display_cols = [
                    "機種",
                    "Type",
                    "Dye",
                    "Lot Number",
                    "NC判定",
                    "PC/Ct値",
                    "PC/SD換算",
                    "PC/判定",
                    "違反ルール",
                    "原因",
                    "CAPA",
                ]
                display_cols = [col for col in display_cols if col in group.columns]
                st.dataframe(group[display_cols], use_container_width=True)
        else:
            st.info("該当期間のデータがありません。")

    # エクセル出力ボタンがクリックされた場合の処理
    if export_excel_button and year and month:
        _, last_day = calendar.monthrange(int(year), int(month))
        start_date = f"{year}-{month.zfill(2)}-01"
        end_date = f"{year}-{month.zfill(2)}-{last_day}"

        with sqlite3.connect("db_folder/qc_data_base.db", timeout=30) as conn:
            query = """
                SELECT
                  f.File_Name,
                  f.Date_Time AS '日時',
                  f.Batch AS 'バッチ',
                  f.Item_Code AS "項目コード",
                  f.Model AS "機種",
                  n.Type AS "Type",
                  n.Dye AS "Dye",
                  n.Lot_Number AS "Lot Number",
                  n.Verdict AS "NC判定",
                  p.Ct AS "PC/Ct値",
                  p.SD_Conversion AS "PC/SD換算",
                  m.judgment AS "PC/判定",
                  m.Violated_rule AS "違反ルール",
                  a.Cause AS "原因",
                  a.CAPA AS "CAPA"
                FROM table_qc_file_info f
                LEFT JOIN table_qc_nc n
                  ON f.File_Name = n.File_Name
                LEFT JOIN table_qc_pc p
                  ON n.File_Name = p.File_Name AND n.Dye = p.Dye
                LEFT JOIN table_qc_multi_rule m
                  ON p.File_Name = m.File_Name AND p.Type = m.Type
                LEFT JOIN table_qc_status_log s
                  ON f.File_Name = s.File_Name
                LEFT JOIN table_qc_act_log a
                  ON p.File_Name = a.File_Name AND p.Type = a.Type
                WHERE
                  strftime('%Y-%m-%d', replace(f.Date_Time, '/', '-')) >= ?
                  AND strftime('%Y-%m-%d', replace(f.Date_Time, '/', '-')) <= ?
                ORDER BY
                  f.Item_Code,
                  n.Type,
                  f.Date_Time DESC
                """
            df = pd.read_sql_query(query, conn, params=(start_date, end_date))
            df = df.drop_duplicates()

        # measurement_item.xlsxを読み込み、項目コードで結合
        df_item = pd.read_excel("data/master/measurement_item.xlsx")
        df = pd.merge(
            df, df_item, left_on="項目コード", right_on="Item_Code", how="left"
        )

        # エクセルファイルを作成
        try:
            filename = create_excel_report(df, year, month)
            st.success(f"エクセルファイル '{filename}' が作成されました。")
        except Exception as e:
            st.error(f"エクセルファイルの作成中にエラーが発生しました: {str(e)}")


# メイン関数を実行
if __name__ == "__main__":
    main()


def display_type_data(conn):
    """Type毎のデータを表示する関数"""
    query = """
    SELECT * FROM table_qc_act_log
    ORDER BY Date_Time DESC
    """
    df = pd.read_sql_query(query, conn)

    if df.empty:
        st.warning("データが存在しません。")
        return

    # Type毎にデータをグループ化
    grouped = df.groupby("Type")

    for type_name, group in grouped:
        st.subheader(f"Type: {type_name}")

        # グラフを描画
        plot_title = f"{type_name}の違反ルール"
        fig = px.line(
            group,
            x="Date_Time",
            y="Violated_rule",
            title=plot_title,
        )
        st.plotly_chart(fig)

        # テーブルを表示
        st.dataframe(group)
